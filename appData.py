from tkinter import *
from tkinter.filedialog import askopenfile
from openpyxl import load_workbook
from tkinter.ttk import Frame,Button, Style
import datetime
import tkinter
import pandas as pd
import re
now= datetime.datetime.today().strftime('%Y-%m-%d_%H%M%S')
def get_cell_value_list(sheet):
    return([[cell.value for cell in row] for row in sheet])
def label_race(row):
   if (row['isCancelFollow'] == True) and (row['isQualifiedContact'] == True) :
      return True
   else:
      return False
def dummy_data(label):
    
    file = askopenfile(mode ='r', filetypes =[('Excel Files', '*.xlsx *.xlsm *.sxc *.ods *.csv *.tsv')]) # To open the file that you want. 
    #' mode='r' ' is to tell the filedialog to read the file
    # 'filetypes=[()]' is to filter the files shown as only Excel files
    wb = load_workbook(filename = file.name) # Load into openpyxl
     # Load into openpyxl
    sheet=wb[wb.sheetnames[0]]
    max_row = sheet.max_row            
    HS=get_cell_value_list(sheet['A1:B{}'.format(max_row)])
    df=pd.DataFrame(HS)
    new_header = df.iloc[0] #grab the first row for the header
    df = df[1:] #take the data less the header row
    df.columns = new_header
    
    if ('Data' in df) and ('Student Code' in df):
        df = df.set_index(['Student Code'])
        dummy=pd.get_dummies(df['Data'].explode()).groupby(level=0).sum()
        file_name=('dummy_'+str(now)).replace(" ", "_")
        print(file_name)
        print(dummy)
        dummy.to_excel('SaveFile/{}.xlsx'.format(file_name))
        label["text"] ="Chúc mừng bạn đã chuyển đổi thành công"
        
    else:
         label["text"] ="Không thể xuất fife được vui lòng kiểm tra lại tên cột"
def pivot_report(label):
    file = askopenfile(mode ='r', filetypes =[('Excel Files', '*.xlsx *.xlsm *.sxc *.ods *.csv *.tsv')]) # To open the file that you want. 
    #' mode='r' ' is to tell the filedialog to read the file
    # 'filetypes=[()]' is to filter the files shown as only Excel files
    wb = load_workbook(filename = file.name) # Load into openpyxl
     # Load into openpyxl
    sheet=wb[wb.sheetnames[0]]
    max_row = sheet.max_row            
    HS=get_cell_value_list(sheet['A1:K{}'.format(max_row)])
    df=pd.DataFrame(HS)
    new_header = df.iloc[0] #grab the first row for the header
    df = df[1:] #take the data less the header row
    df.columns = new_header
    print(df)
    check_columns=['Full Name', 'Enroll Grade: Grade Name', 'Mobile (+84)',
       'Lead Owner: Full Name', 'isCancelFollow', 'isContact',
       'Pre-qualified Contact', 'isQualifiedContact', 'isEnquiry',
       'isProspect', 'isRegistered']
    if check_columns==df.columns.tolist():
        file_name=('pivot_report'+str(now)).replace(" ", "_")
        file_name_2=('raw_data'+str(now)).replace(" ", "_")
        print(file_name)
        df['Cancel_Quali'] = df.apply(label_race, axis=1)
        unpivot=pd.melt(df, id_vars=['Lead Owner: Full Name','Full Name','Enroll Grade: Grade Name','Mobile (+84)'])
        data=unpivot[unpivot.value!=False]
        data=data.rename(columns={0:'varibale'})
        pivot=data.filter(['Lead Owner: Full Name','varibale','value'], axis=1)
        table = pd.pivot_table(pivot, values='value', index='Lead Owner: Full Name',columns='varibale', aggfunc="count")
        table['Cancel_NotQuali']=table['isCancelFollow']-table['Cancel_Quali']
        table.to_excel('SaveFile/{}.xlsx'.format(file_name))
        data.to_excel('SaveFile/{}.xlsx'.format(file_name_2))
        label["text"] ="Chúc mừng bạn đã chuyển đổi thành công"
        
    else:
        label["text"] ="Không thể xuất fife được vui lòng kiểm tra lại tên cột"
def split_name_three(label):
    file = askopenfile(mode ='r', filetypes =[('Excel Files', '*.xlsx *.xlsm *.sxc *.ods *.csv *.tsv')]) # To open the file that you want. 
    #' mode='r' ' is to tell the filedialog to read the file
    # 'filetypes=[()]' is to filter the files shown as only Excel files

    wb = load_workbook(filename = file.name) # Load into openpyxl
    sheet=wb[wb.sheetnames[0]]
    max_row = sheet.max_row            
    HS=get_cell_value_list(sheet['A1:A{}'.format(max_row)])    
    if HS[0][0]=='Full Name':
        file_name=('Three_name_'+str(now)).replace(" ", "_")
        HS.pop(0)
        last_name=[]
        first_name=[]
        midle_name=[]
        first, middle,last=["","",""]
        for i in HS:
           print(len(i[0].split()))
           if (len(i[0].split())>1):
             first, *middle,last = i[0].split()
             last_name.append((last))
             first_name.append(first)
             midle_name.append(" ".join(middle))    
           else:
              last_name.append(i[0])
              first_name.append("")
              midle_name.append("")
        final = pd.DataFrame(
           {'First': first_name,
            'Midle': midle_name,
            'Last': last_name
                  })
        print(file_name)
        final.to_excel('SaveFile/{}.xlsx'.format(file_name))
        label["text"] ="Chúc mừng bạn đã chuyển đổi thành công"
    else:
        label["text"] ="Không thể xuất fife được vui lòng kiểm tra lại tên cột"
def split_name_two(label):
    file = askopenfile(mode ='r', filetypes =[('Excel Files', '*.xlsx *.xlsm *.sxc *.ods *.csv *.tsv')]) # To open the file that you want. 
    #' mode='r' ' is to tell the filedialog to read the file
    # 'filetypes=[()]' is to filter the files shown as only Excel files

    wb = load_workbook(filename = file.name) # Load into openpyxl
    sheet=wb[wb.sheetnames[0]]
    max_row = sheet.max_row            
    HS=get_cell_value_list(sheet['A1:A{}'.format(max_row)])    
    if HS[0][0]=='Full Name':
        file_name=('Two_name_'+str(now)).replace(" ", "_")
        HS.pop(0)
        last_name=[]
        first_name=[]
        first, middle,last=["","",""]
        for i in HS:
           print(len(i[0].split()))
           if (len(i[0].split())>1):
             *first, last = i[0].split()
             last_name.append(last)
             first_name.append(" ".join(first)) 
           else:
              last_name.append(i[0])
              first_name.append("")
             
        final = pd.DataFrame(
           {'First': first_name,
            'Last': last_name
                  })
        print(file_name)
        final.to_excel('SaveFile/{}.xlsx'.format(file_name))
        label["text"] ="Chúc mừng bạn đã chuyển đổi thành công"
    else:
        label["text"] ="Không thể xuất fife được vui lòng kiểm tra lại tên cột"
def group_siblings(label):
    file = askopenfile(mode ='r', filetypes =[('Excel Files', '*.xlsx *.xlsm *.sxc *.ods *.csv *.tsv')]) # To open the file that you want. 
    #' mode='r' ' is to tell the filedialog to read the file
    # 'filetypes=[()]' is to filter the files shown as only Excel files

    wb = load_workbook(filename = file.name) # Load into openpyxl
    sheet=wb[wb.sheetnames[0]]
    max_row = sheet.max_row            
    HS=get_cell_value_list(sheet['A1:D{}'.format(max_row)])
    df=pd.DataFrame(HS)
    new_header = df.iloc[0] #grab the first row for the header
    df = df[1:] #take the data less the header row
    df.columns = new_header
    if ('Family No' in df) and ('Student Code' in df)and ('Student Name' in df)and ('Candidate Id' in df):

        list_out=df.values.tolist()
        total=[]
        for i,row in enumerate(list_out):
             val=((str(row[3])+","+str(row[1])+","+row[2]).split(','))
             list_out[i].append(val)
        for i in list_out:
            total.append([i[0]])
        for i,row in enumerate(total):
            total[i].append(list_out[i][4])
        tmp_dict = {}

        for first, second in total:
             if first not in tmp_dict:
                 tmp_dict[first] = []
             tmp_dict[first].append(second)
        new_list = []

        for key, val in tmp_dict.items():
            new_list.append([key] + val)

        final=pd.DataFrame(new_list)
        final.rename(columns={0:'Family No'}, inplace=True)
        for i in range(1,len(final.loc[1])):
          final.rename(columns={i:'Student No.{}'.format(i)}, inplace=True)
        for i in range(1,len(final.loc[1])):
            mask = final['Student No.{}'.format(i)].notnull()
            final.loc[mask, 'Student No.{}'.format(i)] = [','.join(map(str, x)) for x in final.loc[mask, 'Student No.{}'.format(i)]]
            file_name=('Sibling'+str(now)).replace(" ", "_")
        print(file_name)
        
        final.to_excel('SaveFile/{}.xlsx'.format(file_name))
        label["text"] ="Chúc mừng bạn đã chuyển đổi thành công"
        
    else:
        label["text"] ="Không thể xuất fife được vui lòng kiểm tra lại tên cột"

class Example(Frame):
    def __init__(self,parent):
        Frame.__init__(self,parent)
        self.parent=parent
        self.initUI()

    def initUI(self):
        label=Label(root, text='Hãy nhập file và đợi kết quả',font=('Helvetica 13 bold'))
        label.place(x=500,y=550)
        Label(root, text='Các bạn hãy xem file hướng dẫn trước để làm không bị sai nha và nhớ lưu ý nhập đúng template',font=('Helvetica 13 bold')).place(x=250,y=150)
        Label(root, text='Ở mục trả kết quả nếu đợi quá lâu không thay đổi dòng chữ ban đầu thì có thể do file input quá nặng -->đợi lâu. Hoặc là code bị lỗi -->không bao giờ ra kết quả được :v',font=('Helvetica 13 bold')).place(x=13,y=350)
        self.parent.title("Excel Processing")
        self.style=Style()
        self.style.theme_use("default")
        self.pack(fill=BOTH, expand=1)
        btn = Button(root, text ='Dummy Data', command =lambda:dummy_data(label))
        btn.place(x=100,y=250)
        btn = Button(root, text ='Group By Silbing', command =lambda:group_siblings(label))
        btn.place(x=350,y=250)
        btn = Button(root, text ='Split Full Name Three', command =lambda:split_name_three(label))
        btn.place(x=600,y=250)
        btn = Button(root, text ='Split Full Name Two', command =lambda:split_name_two(label))
        btn.place(x=850,y=250)
        btn = Button(root, text ='Report Table', command =lambda:pivot_report(label))
        btn.place(x=1100,y=250)
        
root = Tk()
w, h = root.winfo_screenwidth(), root.winfo_screenheight()
root.geometry("%dx%d+0+0" % (w, h))
app=Example(root)
root.mainloop()



